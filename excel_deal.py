 
import openpyxl
from structs import ProductStorie
from structs import TaskStruct 
import datetime
class excel_deal(object):
    def __init__(self,sheet_name):
        self.column_indices = []
        self.row_index_=3
        self.row_max =2
        self.columns_to_copy = ["项目名称",'需求', 
        '描述',
         '输出成果/验收标准', 
         '执行人',
         '优先级',
         "要求完成时间",
         "用户故事/任务",
         "初始规模",
         "计划完成时间"]
        self.GetExcelInfo(sheet_name)

    def GetExcelInfo(self,source_file,sheet_name):
        # 打开原始Excel文件 
        wb1 = openpyxl.load_workbook(source_file)

        # 选择要复制的sheet
        self.sheet_source = wb1[sheet_name]
        # 复制指定列的数据
        
        for column in self.columns_to_copy:
            for cell in self.sheet_source[2]:
                if cell.value == "需求负责人":
                    continue
                if cell.value == column or str(cell.value).find(column)==0:
                    self.column_indices.append(cell.column)                             
        self.row_max = self.sheet_source.max_row
        
    def SaveExcelInfo(self):
        # 创建新的Excel文件
        new_file = '新文件.xlsx'
        wb2 = openpyxl.Workbook()
        sheet2 = wb2.active
        column_index=1
        for index in self.column_indices:    
            print("index:%d ----------------------" % (index))
            for row in range(2, self.row_max + 1):
                sheet2.cell(row=row, column=column_index, value=self.sheet_source.cell(row=row, column=index).value) 
                print("Sheet2 row:%d colunm:%d data:%s" % (row,column_index,sheet2.cell(row,index).value))
            column_index = column_index+1
        # 保存新的Excel文件
        wb2.save(new_file)

    def GetNextRow(self):
        pro = ProductStorie()
        taskStruct = TaskStruct()
        if self.row_index_>self.row_max:
            return False,pro,taskStruct
        else:
#            print("GetOneRow index:%d" % self.row_index_)
            pro.title = self.sheet_source.cell(
                row=self.row_index_, column=self.GetColumnsIndx("需求")).value      
            pro.spec = self.sheet_source.cell(
                row=self.row_index_, column=self.GetColumnsIndx("描述")).value
            pro.category = "feature"     
            pro.product = self.sheet_source.cell(
                row=self.row_index_, column=self.GetColumnsIndx("项目名称")).value
            pro.pri = self.sheet_source.cell(
                row=self.row_index_, column=self.GetColumnsIndx("优先级")).value  
            pro.verify =self.sheet_source.cell(
                row=self.row_index_, column=self.GetColumnsIndx("输出成果/验收标准")).value 

            taskStruct.name=self.sheet_source.cell(
                row=self.row_index_, 
                column=self.GetColumnsIndx("用户故事/任务")).value
#            print("GetOneRow index:%d pro.title:%s taskStruct.name:%s" 
#                % (self.row_index_,pro.title,taskStruct.name))

   #         print("GetOneRow index:%d column:%d taskStruct.name:%s" 
 #               % (self.row_index_,self.GetColumnsIndx("用户故事/任务"),taskStruct.name))
            taskStruct.type="devel"  
            taskStruct.pri=3
            taskStruct.estimate=self.sheet_source.cell(
                row=self.row_index_, 
                column=self.GetColumnsIndx('初始规模')).value
            try:
                taskStruct.deadline=self.sheet_source.cell(
                    row=self.row_index_, 
                    column=self.GetColumnsIndx('计划完成时间')).value.strftime("%Y-%m-%d")
            except: 
                next_month = datetime.datetime.now().replace(day=28) + datetime.timedelta(days=4)
                taskStruct.deadline= next_month - datetime.timedelta(days=next_month.day) 
            taskStruct.assignedTo=self.sheet_source.cell(
                row=self.row_index_, 
                column=self.GetColumnsIndx('执行人')).value 
            taskStruct.estStarted=datetime.datetime.now().strftime("%Y-%m-%d")   
            self.row_index_ = self.row_index_+1
            return True,pro,taskStruct

    def GetColumnsIndx(self,name):
            return self.column_indices[
                    self.columns_to_copy.index(name)]
   
    def GetExcelRowsNum(self):
        return self.row_max

if __name__ == "__main__":
    exlce_ = excel_deal("tmp.xlsx","月度需求计划-5月") 
    exlce_.GetNextRow()
   
    
    